import openpyxl
import re
import os

EXCEL_PATH = "/Applications/SAFAWALA MASTERPLAN/INVENTORY DETAILS/MALAS/MALAS.xlsx"

def shorten_color(col):
    if not col:
        return ""
    col_str = str(col).strip()
    low = col_str.lower()
    
    keywords = [
        ('maroon', 'Maroon'),
        ('crimson', 'Crimson'),
        ('ruby', 'Ruby'),
        ('red', 'Red'),
        ('pink', 'Pink'),
        ('blush', 'Pink'),
        ('peach', 'Peach'),
        ('emerald', 'Green'),
        ('green', 'Green'),
        ('turquoise', 'Turquoise'),
        ('silver', 'Silver'),
        ('gold', 'Gold'),
        ('champagne', 'Gold'),
        ('white', 'White'),
        ('ivory', 'White'),
        ('cream', 'White'),
        ('mixed', 'Mixed')
    ]
    
    earliest_idx = float('inf')
    matched_val = None
    
    for key, val in keywords:
        idx = low.find(key)
        if idx != -1 and idx < earliest_idx:
            earliest_idx = idx
            matched_val = val
            
    if matched_val:
        return matched_val
        
    match = re.search(r'[a-zA-Z]+', col_str)
    if match:
        word = match.group(0)
        return word.capitalize()
    return col_str

def shorten_size(size):
    if not size:
        return "Standard"
    size_str = str(size).strip()
    low = size_str.lower()
    
    if 'adjustable' in low:
        return 'Adjustable'
    if 'standard' in low:
        return 'Standard'
        
    match = re.search(r'[a-zA-Z]+', size_str)
    if match:
        word = match.group(0)
        return word.capitalize()
    return 'Standard'

def get_ultra_short_material(mat):
    if not mat:
        return "Beads"
    mat_str = str(mat).strip()
    low = mat_str.lower()
    
    keywords = [
        ('kundan', 'Kundan'),
        ('polki', 'Polki'),
        ('pearl', 'Pearls'),
        ('zircon', 'Zircon'),
        ('cz', 'Zircon'),
        ('crystal', 'Crystals'),
        ('stone', 'Beads'),
        ('bead', 'Beads'),
        ('metal', 'Metal'),
        ('alloy', 'Metal')
    ]
    
    earliest_idx = float('inf')
    matched_val = None
    
    for key, val in keywords:
        idx = low.find(key)
        if idx != -1 and idx < earliest_idx:
            earliest_idx = idx
            matched_val = val
            
    if matched_val:
        return matched_val
        
    match = re.search(r'[a-zA-Z]+', mat_str)
    if match:
        word = match.group(0)
        return word.capitalize()
    return "Beads"

def run():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    print(f"Sheet title: {ws.title}")
    print(f"Original images count: {len(ws._images)}")
    
    updated_count = 0
    
    # Data starts from row 4 in 1-based index (S.No 1 is in row 4)
    # Row 1: Title, Row 2: Register, Row 3: Headers
    for row in range(4, ws.max_row + 1):
        sno = ws.cell(row=row, column=1).value
        # Check if it's a valid product row (S.No is a number)
        if not isinstance(sno, (int, float)):
            continue
            
        cell = ws.cell(row=row, column=3) # Column C is "Colour | Size | Material"
        original_val = cell.value
        if not original_val:
            continue
            
        parts = [p.strip() for p in str(original_val).split('|')]
        color_str = parts[0] if len(parts) > 0 else ""
        size_str = parts[1] if len(parts) > 1 else ""
        material_str = parts[2] if len(parts) > 2 else ""
        
        final_color = shorten_color(color_str)
        final_size = shorten_size(size_str)
        final_material = get_ultra_short_material(material_str)
        
        cleaned_val = f"{final_color} | {final_size} | {final_material}"
        
        if original_val != cleaned_val:
            print(f"Row {row} (S.No {sno}):")
            print(f"  Old: {original_val}")
            print(f"  New: {cleaned_val}")
            cell.value = cleaned_val
            updated_count += 1
            
    if updated_count > 0:
        print(f"\nSaving changes to workbook...")
        wb.save(EXCEL_PATH)
        print(f"Workbook saved successfully. Images count: {len(ws._images)}")
    else:
        print("\nNo rows required updating.")
        
run()
