import openpyxl, os

BASE = "/Users/rahulmedhe/Downloads/INVENTORY SETUP"

# Each category: (folder, filename, [(Product Name, Material, Size, Colour), ...])
CATEGORIES = [

("TALWAR", "TALWAR PRICING.xlsx", [
    ("Royal Sword",        "Metal", "Large", "Brown"),
    ("Dragon Sword",       "Metal", "Large", "White"),
    ("Mughal Sword",       "Metal", "Large", "White"),
    ("Velvet Sword",       "Metal", "Large", "Black"),
    ("Bridal Sword",       "Metal", "Large", "Red"),
    ("Rajput Sword",       "Metal", "Large", "Red"),
    ("Teal Turban Pin",    "Metal", "Free",  "Teal"),
    ("Chrome Sword",       "Metal", "Large", "Black"),
    ("Pearl Sword",        "Metal", "Large", "White"),
    ("Filigree Sword",     "Metal", "Large", "Black"),
    ("Golden Sword",       "Metal", "Large", "White"),
    ("Teak Sword",         "Metal", "Large", "Brown"),
    ("Silver Sword",       "Metal", "Large", "Silver"),
    ("Zari Sword",         "Metal", "Large", "Red"),
    ("Steel Sword",        "Metal", "Large", "Black"),
    ("Ruby Turban Pin",    "Metal", "Free",  "Ruby"),
    ("Heritage Sword",     "Metal", "Large", "Brown"),
    ("Ruby Turban Pin II", "Metal", "Free",  "Ruby"),
    ("Brass Sword",        "Metal", "Large", "Brown"),
    ("Classic Sword",      "Metal", "Large", "Brown"),
    ("Lal Sword",          "Metal", "Large", "Red"),
    ("Ivory Sword",        "Metal", "Large", "White"),
    ("Dragon Sword II",    "Metal", "Large", "Silver"),
    ("Shadow Sword",       "Metal", "Large", "Black"),
    ("Guard Sword",        "Metal", "Large", "Black"),
    ("Sandal Sword",       "Metal", "Large", "Brown"),
    ("Maroon Sword",       "Metal", "Large", "Maroon"),
]),

("BELT", "BELT PRICING.xlsx", [
    ("Agni Belt",  "Brocade", "Free", "Crimson"),
    ("Surya Belt", "Brocade", "Free", "Crimson"),
    ("Ratna Belt", "Brocade", "Free", "Champagne"),
]),

("WAIST", "WAIST PRICING.xlsx", [
    ("Kesar Sash", "Kota", "Free", "Peach"),
    ("Zari Sash",  "Kota", "Free", "Peach"),
]),

("SCARF", "SCARF PRICING.xlsx", [
    ("Kesar Scarf", "Kota",  "Free", "Peach"),
    ("Zari Scarf",  "Kota",  "Free", "Peach"),
    ("Hawa Scarf",  "Satin", "Free", "Chocolate"),
    ("Bahar Scarf", "Satin", "Free", "Red"),
    ("Neela Scarf", "Satin", "Free", "Blue"),
    ("Shahi Scarf", "Satin", "Free", "Teal"),
]),

("KATAR", "KATAR PRICING.xlsx", [
    ("Tej Dagger",   "Metal",  "Small",  "Gold"),
    ("Nag Dagger",   "Metal",  "Medium", "Gold"),
    ("Veer Dagger",  "Metal",  "Large",  "Gold"),
    ("Arjun Dagger", "Metal",  "Set",    "Gold"),
    ("Kundan Dagger","Velvet", "Medium", "Maroon"),
    ("Chandi Dagger","Metal",  "Medium", "Silver"),
    ("Shahi Dagger", "Metal",  "Large",  "Silver"),
]),

("FEATHER", "FEATHER PRICING.xlsx", [
    ("Hansa Feather", "Feather", "Small",  "White"),
    ("Agni Feather",  "Feather", "Medium", "Saffron"),
    ("Mukut Feather", "Feather", "Medium", "White"),
    ("Pearl Feather", "Feather", "Small",  "White"),
    ("Shubh Feather", "Feather", "Medium", "White"),
    ("Ratna Feather", "Feather", "Medium", "White"),
    ("Raja Feather",  "Feather", "Large",  "White"),
    ("Veer Feather",  "Feather", "Medium", "White"),
    ("Tej Feather",   "Feather", "Small",  "White"),
    ("Surya Feather", "Feather", "Large",  "White"),
]),

("NECK", "NECK PRICING.xlsx", [
    ("Kala Cravat",  "Brocade",   "Free", "Black"),
    ("Lal Cravat",   "Brocade",   "Free", "Red"),
    ("Rani Cravat",  "Brocade",   "Free", "Crimson"),
    ("Neeli Cravat", "Jacquard",  "Free", "Navy"),
    ("Bahar Cravat", "Silk",      "Free", "Blue"),
    ("Gulab Cravat", "Jacquard",  "Free", "Peach"),
    ("Raja Cravat",  "Silk",      "Free", "Maroon"),
    ("Shahi Cravat", "Polyester", "Free", "Navy"),
    ("Neel Cravat",  "Polyester", "Free", "Navy"),
    ("Indra Cravat", "Polyester", "Free", "Navy"),
    ("Tej Cravat",   "Silk",      "Free", "Midnight"),
    ("Krishn Cravat","Satin",     "Free", "Black"),
    ("Surya Cravat", "Jacquard",  "Free", "Navy"),
    ("Shyam Cravat", "Silk",      "Free", "Magenta"),
    ("Veer Cravat",  "Silk",      "Free", "Wine"),
]),

("POCKET BROOCH", "POCKET BROOCH PRICING.xlsx", [
    ("Guitar Brooch",       "Metal", "Free", "Gold"),
    ("Spectacle Brooch",    "Metal", "Free", "Gold"),
    ("Cannon Brooch",       "Metal", "Free", "Gold"),
    ("Ewer Brooch",         "Metal", "Free", "Gold"),
    ("Eagle Brooch",        "Metal", "Free", "Gold"),
    ("Crystal Brooch",      "Metal", "Free", "White"),
    ("Medallion Brooch",    "Metal", "Free", "Gold"),
    ("Ship Brooch",         "Metal", "Free", "Gold"),
    ("Floral Brooch",       "Metal", "Free", "Gold"),
    ("Crown Brooch",        "Metal", "Free", "Gold"),
    ("Pearl Crown Brooch",  "Metal", "Free", "White"),
    ("Tiara Brooch",        "Metal", "Free", "White"),
    ("Shield Brooch",       "Metal", "Free", "Silver"),
    ("Hawk Brooch",         "Metal", "Free", "Silver"),
    ("Moustache Brooch",    "Metal", "Free", "Gold"),
    ("Phoenix Brooch",      "Metal", "Free", "Gold"),
    ("Axe Brooch",          "Metal", "Free", "Gold"),
    ("Palace Brooch",       "Metal", "Free", "Black"),
    ("Pearl Brooch",        "Metal", "Free", "White"),
    ("Kundan Brooch",       "Metal", "Free", "White"),
    ("Pearl Chain Brooch",  "Metal", "Free", "Gold"),
    ("Dragon Brooch",       "Metal", "Free", "Silver"),
    ("Betiwale Brooch",     "Metal", "Free", "Red"),
    ("Trishul Brooch",      "Metal", "Free", "Red"),
    ("Royal Brooch",        "Metal", "Free", "Green"),
    ("Crest Brooch",        "Metal", "Free", "Gold"),
    ("Clan Brooch",         "Metal", "Free", "Red"),
    ("Garuda Brooch",       "Metal", "Free", "Red"),
    ("Sun Brooch",          "Metal", "Free", "Gold"),
    ("Surya Brooch",        "Metal", "Free", "Gold"),
    ("Regal Brooch",        "Metal", "Free", "Green"),
]),

]

# ---- GROOM UPDATED ----
CATEGORIES.append(("GROOM UPDATED", "GROOM UPDATED PRICING.xlsx", [
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "White"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "White (Shimmer)"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Off-White"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Light Pink"),
    ("Groom Safa - Wedding Turban", "Velvet",  "Free Size", "Pink"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Peach"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Cream/Gold"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Ivory"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Cream (Gold Check)"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Yellow/Cream"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Peach"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Magenta/Dark Pink (Striped)"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Fuchsia/Hot Pink"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Sky Blue"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Yellow"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Baby Pink"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Mint Green (Gold Check)"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Magenta (Gold Check)"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Lime Green"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Magenta"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Baby Pink"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Teal/Aqua"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Peach"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Sage Green"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Gold/Champagne"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Blush"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Cream/Ivory"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Dusty Rose/Mauve"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Blush/Peach"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Peach/Gold"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Off-White"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "Steel Blue/Grey"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "White (Gold Lace Border)"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "White (Gold Buti)"),
    ("Groom Safa - Wedding Turban", "Velvet",  "Free Size", "Ivory/White"),
    ("Groom Safa - Wedding Turban", "Silk",    "Free Size", "White (Sheer)"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Baby Pink"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Grey (Gold Buti)"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Peach (Gold Buti)"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "Lime Green (Gold Buti)"),
    ("Groom Safa - Wedding Turban", "Brocade", "Free Size", "White (Gold Buti)"),
    ("Groom Safa - Wedding Turban", "Satin",   "Free Size", "Red/Maroon"),
]))

# ---- VELCRO ----
CATEGORIES.append(("VELCRO", "VELCRO PRICING.xlsx", [
    ("Bandhani Turban",       "Bandhani",   "Free", "Multicolor"),
    ("Kesari Turban",         "Cotton",     "Free", "Orange"),
    ("Rainbow Turban",        "Georgette",  "Free", "Multicolor"),
    ("Sona Turban",           "Silk",       "Free", "Yellow"),
    ("Ivory Turban",          "Silk",       "Free", "Ivory"),
    ("Blush Turban",          "Chiffon",    "Free", "Blush"),
    ("Rose Turban",           "Brocade",    "Free", "Pink"),
    ("Brocade Turban",        "Brocade",    "Free", "Pink"),
    ("Salmon Turban",         "Cotton",     "Free", "Salmon"),
    ("Peach Turban",          "Tissue",     "Free", "Peach"),
    ("Bandhani Pink Turban",  "Bandhani",   "Free", "Pink"),
    ("Shimmer Turban",        "Tissue",     "Free", "Pink"),
    ("Chikankari Turban",     "Chikankari", "Free", "Pink"),
    ("Silk Turban",           "Silk",       "Free", "Pink"),
    ("Lal Turban",            "Tissue",     "Free", "Red"),
    ("Kota Turban",           "Kota",       "Free", "Maroon"),
    ("Stripe Turban",         "Kota",       "Free", "Red"),
    ("Tomato Turban",         "Satin",      "Free", "Red"),
    ("Peach Check Turban",    "Cotton",     "Free", "Peach"),
    ("Apricot Turban",        "Silk",       "Free", "Peach"),
    ("Dusty Rose Turban",     "Cotton",     "Free", "Salmon"),
    ("Champagne Turban",      "Georgette",  "Free", "Champagne"),
    ("Pearl Turban",          "Silk",       "Free", "White"),
    ("Floral Orange Turban",  "Cotton",     "Free", "Orange"),
    ("Ivory Floral Turban",   "Georgette",  "Free", "White"),
    ("Coral Turban",          "Cotton",     "Free", "Coral"),
    ("Garden Turban",         "Georgette",  "Free", "Gold"),
    ("Ikat Turban",           "Cotton",     "Free", "White"),
    ("Magenta Turban",        "Satin",      "Free", "Magenta"),
    ("Rani Silk Turban",      "Silk",       "Free", "Magenta"),
    ("Gold Shimmer Turban",   "Tissue",     "Free", "White"),
    ("Gold Check Turban",     "Tissue",     "Free", "White"),
    ("Multicolor Turban",     "Cotton",     "Free", "Multicolor"),
    ("Leheria Turban",        "Cotton",     "Free", "Orange"),
    ("Mint Turban",           "Georgette",  "Free", "Mint"),
    ("Velvet Turban",         "Velvet",     "Free", "Maroon"),
    ("Zardozi Turban",        "Satin",      "Free", "Red"),
    ("Copper Turban",         "Silk",       "Free", "Maroon"),
    ("Silver Turban",         "Satin",      "Free", "Silver"),
    ("Gold Satin Turban",     "Satin",      "Free", "Gold"),
    ("Rani Turban",           "Satin",      "Free", "Magenta"),
    ("Purple Turban",         "Satin",      "Free", "Purple"),
    ("Maroon Turban",         "Satin",      "Free", "Maroon"),
    ("Scarlet Turban",        "Satin",      "Free", "Red"),
    ("Blue Turban",           "Satin",      "Free", "Blue"),
    ("Teal Turban",           "Satin",      "Free", "Teal"),
    ("Yellow Turban",         "Satin",      "Free", "Gold"),
    ("Party Turban",          "Cotton",     "Free", "Multicolor"),
    ("Tricolor Turban",       "Cotton",     "Free", "Tricolor"),
    ("Mini Turban",           "Satin",      "Free", "Pink"),
]))


# ---- MALAS ----
CATEGORIES.append(("MALAS", "MALAS PRICING.xlsx", [
    ("CZ Diamond Groom Mala",           "Metal/CZ",          "Long", "Silver/White"),
    ("3-Layer CZ Kundan Mala",          "Metal/CZ",          "Long", "White/Silver"),
    ("Pearl Rudraksha Kundan Mala",     "Pearl/Metal",       "Long", "White/Maroon"),
    ("2-Layer Rudraksha Crystal Mala",  "Rudraksha/Metal",   "Long", "Maroon/White"),
    ("Kundan Triangle Pink Mala",       "Metal/Kundan",      "Long", "White/Pink"),
    ("4-Layer Ruby CZ Mala",            "Metal/CZ",          "Long", "Pink/Ruby"),
    ("Rudraksha Pearl Gold Charm Mala", "Rudraksha/Pearl",   "Long", "Maroon/White"),
    ("3-Layer CZ Diamond Mala",         "Metal/CZ",          "Long", "White/Silver"),
    ("2-Layer Maroon Pearl Mala",       "Pearl/Metal",       "Long", "Maroon/White"),
    ("5-Layer Pearl Kundan Mala",       "Pearl/Metal",       "Long", "White"),
    ("3-Layer Kundan Black Stone Mala", "Metal/Kundan",      "Long", "White/Black"),
    ("4-Layer Pearl Kundan Dangler",    "Pearl/Metal",       "Long", "White"),
    ("3-Layer Emerald Bead CZ Mala",    "Metal/Beads",       "Long", "Green/White"),
    ("3-Layer Kundan Rice Pearl Mala",  "Pearl/Metal",       "Long", "White"),
    ("3-Layer Pearl Kundan Triangle",   "Pearl/Metal",       "Long", "White"),
    ("Kundan Teardrop White Pink Mala", "Pearl/Metal",       "Long", "White/Pink"),
    ("3-Layer Pearl Pink Bead Mala",    "Pearl/Metal",       "Long", "White/Pink"),
    ("Kundan Triangle Ruby Mala",       "Metal/Kundan",      "Long", "White/Red"),
    ("Single Kundan Teardrop Mala",     "Metal/Kundan",      "Long", "White"),
    ("Kundan Mirror Single Mala",       "Metal/Kundan",      "Long", "White/Mirror"),
    ("3-Layer CZ Peach Enamel Mala",    "Metal/CZ",          "Long", "White/Peach"),
    ("2-Layer CZ Floral Pendant Mala",  "Metal/CZ",          "Long", "White/Gold"),
    ("2-Layer CZ Floral Mala",          "Metal/CZ",          "Long", "White"),
    ("5-Layer Maroon Pearl Mala",       "Pearl/Metal",       "Long", "Maroon/White"),
    ("3-Layer Pearl Cascade Mala",      "Pearl/Metal",       "Long", "White"),
    ("4-Layer Pearl Ruby Oval Mala",    "Pearl/Metal",       "Long", "White/Maroon"),
    ("3-Layer CZ Pearl Maroon Mala",    "Pearl/Metal",       "Long", "White/Maroon"),
    ("4-Layer Maroon Peacock Mala",     "Metal/Beads",       "Long", "Maroon"),
    ("Cascade Pearl Maroon Kundan",     "Pearl/Metal",       "Long", "White/Maroon"),
    ("3-Layer Pearl Kundan Red Tassel", "Pearl/Metal",       "Long", "White"),
    ("2-Layer CZ Pink Stone Mala",      "Metal/CZ",          "Long", "White/Pink"),
    ("3-Layer Rice Pearl Flower Mala",  "Pearl/Metal",       "Long", "White"),
    ("4-Layer Pearl Flower Mala",       "Pearl/Metal",       "Long", "White"),
    ("Pearl Kundan Groom Mala",         "Pearl/Metal",       "Long", "White"),
    ("3-Layer Polki Oval Mala",         "Metal/Kundan",      "Long", "White/Gold"),
    ("3-Layer CZ Teardrop Mala",        "Metal/CZ",          "Long", "White"),
    ("3-Layer CZ Pink Fan Mala",        "Metal/CZ",          "Long", "White/Pink"),
    ("5-Layer Pearl Green Bead Mala",   "Pearl/Metal",       "Long", "White/Green"),
    ("4-Layer Pearl CZ Mala",           "Pearl/Metal",       "Long", "White"),
    ("4-Layer Pearl Kundan Mala",       "Pearl/Metal",       "Long", "White"),
    ("3-Layer Green Bead CZ Mala",      "Metal/Beads",       "Long", "Green/White"),
    ("3-Layer Pearl Kundan Circle",     "Pearl/Metal",       "Long", "White"),
    ("3-Layer Polki Kundan Mala",       "Metal/Kundan",      "Long", "White/Gold"),
    ("3-Layer Pearl Dark Red Mala",     "Pearl/Metal",       "Long", "White"),
    ("3-Layer Kundan Green Pendant",    "Metal/Kundan",      "Long", "White/Green"),
    ("4-Layer Maroon CZ Mala",          "Metal/Beads",       "Long", "Maroon/White"),
    ("4-Layer Maroon Kundan Mala",      "Metal/Beads",       "Long", "Maroon/White"),
    ("2-Layer Maroon CZ Flower Mala",   "Metal/Beads",       "Long", "Maroon/White"),
    ("3-Layer Rect CZ Pearl Mala",      "Pearl/Metal",       "Long", "White"),
    ("4-Layer Champagne Pearl Mala",    "Pearl/Metal",       "Long", "White/Champagne"),
    ("4-Layer Pearl Green Bead Mala",   "Pearl/Metal",       "Long", "White/Green"),
    ("2-Layer Kundan Teardrop Mala",    "Metal/Kundan",      "Long", "White"),
    ("2-Layer Pearl Polki Triangle",    "Pearl/Metal",       "Long", "White/Gold"),
    ("2-Layer CZ Red Stone Mala",       "Metal/CZ",          "Long", "White/Red"),
    ("Cascade Pearl Maroon Fan Mala",   "Pearl/Metal",       "Long", "White/Maroon"),
    ("Single Kundan Heart Mala",        "Metal/Kundan",      "Long", "White"),
    ("3-Layer Green Bead CZ Flat",      "Metal/Beads",       "Long", "Green/White"),
    ("Single Mixed Kundan Long Mala",   "Pearl/Metal",       "Long", "White"),
    ("3-Layer Pearl CZ Flower Mala",    "Pearl/Metal",       "Long", "White"),
    ("3-Layer Pearl Maroon Kundan",     "Pearl/Metal",       "Long", "White/Maroon"),
    ("4-Layer Rice Round Pearl Mala",   "Pearl/Metal",       "Long", "White"),
    ("3-Layer Pearl Kundan Pink Mala",  "Pearl/Metal",       "Long", "White/Pink"),
    ("3-Layer CZ Oval Pink Mala",       "Metal/CZ",          "Long", "White/Pink"),
    ("3-Layer CZ Square Red Mala",      "Metal/CZ",          "Long", "White/Red"),
    ("3-Layer Maroon CZ Black Mala",    "Metal/Beads",       "Long", "Maroon/White"),
    ("5-Layer Rice Pearl CZ Mala",      "Pearl/Metal",       "Long", "White"),
    ("3-Layer Rice Pearl Kundan Sq",    "Pearl/Metal",       "Long", "White"),
    ("3-Layer Magenta Pearl Mala",      "Pearl/Metal",       "Long", "Pink/White"),
    ("3-Layer Pearl Filigree Mala",     "Pearl/Metal",       "Long", "White"),
    ("5-Layer Pearl Maroon Pendant",    "Pearl/Metal",       "Long", "White/Maroon"),
    ("4-Layer Pearl Gold Ball Mala",    "Pearl/Metal",       "Long", "White/Gold"),
    ("4-Layer Pearl Kundan Pendant",    "Pearl/Metal",       "Long", "White"),
    ("3-Layer CZ Green Stone Mala",     "Pearl/Metal",       "Long", "White/Green"),
    ("3-Layer All-Pink CZ Oval Mala",   "Metal/CZ",          "Long", "White/Pink"),
    ("3-Layer CZ Sky Blue Mala",        "Metal/CZ",          "Long", "White/Sky Blue"),
    ("3-Layer Kundan Ruby Mala",        "Metal/Kundan",      "Long", "White/Red"),
    ("3-Layer Pearl Maroon Dangler",    "Pearl/Metal",       "Long", "White/Maroon"),
    ("3-Layer Pink Kundan Pearl Mala",  "Pearl/Metal",       "Long", "White/Pink"),
    ("3-Layer Pearl Pink Bead Kundan",  "Pearl/Metal",       "Long", "White/Pink"),
    ("3-Layer Pearl Maroon CZ Mala",    "Pearl/Metal",       "Long", "White/Maroon"),
    ("3-Layer Kundan Green Rice Mala",  "Pearl/Metal",       "Long", "White/Green"),
    ("3-Layer Pearl Peach Bead Mala",   "Pearl/Metal",       "Long", "White/Peach"),
    ("3-Layer CZ Oval Blue-White",      "Metal/CZ",          "Long", "White"),
    ("3-Layer Kundan Rice Pearl Green", "Pearl/Metal",       "Long", "White/Green"),
    ("Kundan Teardrop Green Rice Mala", "Pearl/Metal",       "Long", "White/Green"),
    ("Kundan Teardrop Peach Rice Mala", "Pearl/Metal",       "Long", "White/Peach"),
    ("Kundan Green Teardrop Pearl",     "Pearl/Metal",       "Long", "White/Green"),
]))


def update_file(folder, filename, data):
    path = os.path.join(BASE, folder, filename)
    if not os.path.exists(path):
        print(f"  SKIPPED — file not found: {path}")
        return

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    max_col = ws.max_column

    # Row 2 = column headers (2-row header structure)
    ws.cell(row=2, column=max_col+1).value = "Product Name"
    ws.cell(row=2, column=max_col+2).value = "Material"
    ws.cell(row=2, column=max_col+3).value = "Size"
    ws.cell(row=2, column=max_col+4).value = "Colour"

    # Data starts at row 3
    for i, (name, material, size, colour) in enumerate(data):
        row = i + 3
        ws.cell(row=row, column=max_col+1).value = name
        ws.cell(row=row, column=max_col+2).value = material
        ws.cell(row=row, column=max_col+3).value = size
        ws.cell(row=row, column=max_col+4).value = colour

    wb.save(path)
    print(f"  DONE — {filename} ({len(data)} rows updated)")


print("=== Safawala Inventory Update ===\n")
for folder, filename, data in CATEGORIES:
    print(f"Updating {folder}...")
    update_file(folder, filename, data)

print("\nAll done! Open your Excel files to verify.")
