import openpyxl

FILE = "/Users/rahulmedhe/Downloads/INVENTORY SETUP/TALWAR/TALWAR PRICING.xlsx"

# Product data: (Product Name, Material, Size, Colour)
data = [
    ("Heritage Sword", "Wood",   "Large", "Brown"),
    ("Dragon Sword",   "Metal",  "Large", "White"),
    ("Royal Sword",    "Metal",  "Large", "White"),
    ("Velvet Sword",   "Velvet", "Large", "Black"),
    ("Bridal Sword",   "Fabric", "Large", "Red"),
    ("Scarlet Sword",  "Fabric", "Large", "Red"),
    ("Teal Turban Pin","Metal",  "Free",  "Teal"),
    ("Chrome Sword",   "Metal",  "Large", "Black"),
    ("Golden Sword",   "Metal",  "Large", "White"),
    ("Filigree Sword", "Velvet", "Large", "Black"),
    ("Swan Sword",     "Metal",  "Large", "White"),
    ("Teak Sword",     "Metal",  "Large", "Brown"),
    ("Silver Dragon Sword","Metal","Large","Silver"),
    ("Zari Sword",     "Fabric", "Large", "Red"),
    ("Steel Sword",    "Metal",  "Large", "Black"),
    ("Ruby Turban Pin","Metal",  "Free",  "Ruby"),
    ("Brass Sword",    "Metal",  "Large", "Brown"),
    ("Ruby Turban Pin II","Metal","Free", "Ruby"),
    ("Rosewood Sword", "Metal",  "Large", "Brown"),
    ("Classic Sword",  "Metal",  "Large", "Brown"),
    ("Lal Sword",      "Fabric", "Large", "Red"),
    ("Ivory Sword",    "Metal",  "Large", "White"),
    ("Dual Dragon Sword","Metal","Large", "Silver"),
    ("Shadow Sword",   "Velvet", "Large", "Black"),
    ("Guard Sword",    "Metal",  "Large", "Black"),
    ("Sandal Sword",   "Metal",  "Large", "Brown"),
    ("Maroon Sword",   "Velvet", "Large", "Maroon"),
]

wb = openpyxl.load_workbook(FILE)
ws = wb.active

# Find the last used column to add new ones after
max_col = ws.max_column

# Add headers in next available columns
ws.cell(row=1, column=max_col+1).value = "Product Name"
ws.cell(row=1, column=max_col+2).value = "Material"
ws.cell(row=1, column=max_col+3).value = "Size"
ws.cell(row=1, column=max_col+4).value = "Colour"

# Fill data starting from row 2
for i, (name, material, size, colour) in enumerate(data):
    row = i + 2
    ws.cell(row=row, column=max_col+1).value = name
    ws.cell(row=row, column=max_col+2).value = material
    ws.cell(row=row, column=max_col+3).value = size
    ws.cell(row=row, column=max_col+4).value = colour

wb.save(FILE)
print(f"Done! Updated {len(data)} rows in TALWAR PRICING.xlsx")
